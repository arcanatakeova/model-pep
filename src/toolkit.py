"""ARCANA AI — Unified Toolkit Layer.

Wraps all installed free toolkits into clean, reusable utilities
that every ARCANA module can import. No module needs to know which
underlying library is being used.

Capabilities:
- NLP: sentiment analysis, entity extraction, fuzzy matching
- Web: scraping, RSS feeds, user-agent rotation
- Content: markdown conversion, HTML sanitization, slug generation
- Validation: email, URL, phone number
- Media: image resizing, watermarking, format conversion
- Cache: in-memory (TTL/LRU) and disk-based persistent cache
- Logging: structured logging with loguru
- Metrics: Prometheus counters/histograms
- Templates: Jinja2 email/report templates
- Formatting: human-readable numbers, dates, tables
"""

from __future__ import annotations

import hashlib
import logging
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

# ═══════════════════════════════════════════════
# NLP & TEXT PROCESSING
# ═══════════════════════════════════════════════

def sentiment_score(text: str) -> dict[str, float]:
    """Analyze sentiment using TextBlob. Returns polarity (-1 to 1) and subjectivity (0 to 1)."""
    from textblob import TextBlob
    blob = TextBlob(text)
    return {
        "polarity": round(blob.sentiment.polarity, 3),
        "subjectivity": round(blob.sentiment.subjectivity, 3),
        "label": "positive" if blob.sentiment.polarity > 0.1
                 else "negative" if blob.sentiment.polarity < -0.1
                 else "neutral",
    }


def fuzzy_match(s1: str, s2: str) -> float:
    """Fuzzy string similarity ratio (0-1). Uses Levenshtein distance."""
    try:
        import Levenshtein
        return Levenshtein.ratio(s1.lower(), s2.lower())
    except ImportError:
        # Fallback: simple character overlap
        set1, set2 = set(s1.lower()), set(s2.lower())
        if not set1 or not set2:
            return 0.0
        return len(set1 & set2) / len(set1 | set2)


def is_duplicate_text(text1: str, text2: str, threshold: float = 0.85) -> bool:
    """Check if two texts are near-duplicates."""
    return fuzzy_match(text1, text2) >= threshold


def extract_keywords(text: str, top_n: int = 10) -> list[str]:
    """Extract keywords from text using TextBlob noun phrase extraction."""
    from textblob import TextBlob
    blob = TextBlob(text)
    # Get noun phrases and single nouns
    phrases = [str(p).lower() for p in blob.noun_phrases]
    # Deduplicate and rank by frequency
    freq: dict[str, int] = {}
    for p in phrases:
        freq[p] = freq.get(p, 0) + 1
    return sorted(freq.keys(), key=lambda k: -freq[k])[:top_n]


def detect_language(text: str) -> str:
    """Detect text language using TextBlob."""
    try:
        from textblob import TextBlob
        blob = TextBlob(text[:500])
        return str(blob.detect_language())
    except Exception:
        return "en"  # Default to English


# ═══════════════════════════════════════════════
# WEB SCRAPING
# ═══════════════════════════════════════════════

def scrape_html(html: str, selector: str = "", tag: str = "") -> list[str]:
    """Extract text from HTML using BeautifulSoup.

    Args:
        html: Raw HTML string
        selector: CSS class to filter by (e.g., "post-title")
        tag: HTML tag to find (e.g., "h2", "p", "a")
    """
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(html, "lxml")

    if selector:
        elements = soup.find_all(class_=selector)
    elif tag:
        elements = soup.find_all(tag)
    else:
        elements = [soup]

    return [el.get_text(strip=True) for el in elements if el.get_text(strip=True)]


def extract_links(html: str, base_url: str = "") -> list[dict[str, str]]:
    """Extract all links from HTML."""
    from bs4 import BeautifulSoup
    from furl import furl
    soup = BeautifulSoup(html, "lxml")
    links = []
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if base_url and href.startswith("/"):
            href = str(furl(base_url).set(path=href))
        links.append({
            "url": href,
            "text": a.get_text(strip=True),
        })
    return links


def parse_rss_feed(xml_content: str) -> list[dict[str, str]]:
    """Parse an RSS/Atom feed and return entries."""
    import atoma
    try:
        feed = atoma.parse_rss_bytes(xml_content.encode())
        return [
            {
                "title": item.title or "",
                "link": item.link or "",
                "description": (item.description or "")[:500],
                "published": str(item.pub_date) if item.pub_date else "",
            }
            for item in feed.items
        ]
    except Exception:
        try:
            feed = atoma.parse_atom_bytes(xml_content.encode())
            return [
                {
                    "title": entry.title.value if entry.title else "",
                    "link": entry.links[0].href if entry.links else "",
                    "description": (entry.summary.value if entry.summary else "")[:500],
                    "published": str(entry.updated) if entry.updated else "",
                }
                for entry in feed.entries
            ]
        except Exception:
            return []


def random_user_agent() -> str:
    """Get a random browser user-agent string."""
    try:
        from fake_useragent import UserAgent
        return UserAgent().random
    except Exception:
        return "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"


# ═══════════════════════════════════════════════
# CONTENT & MARKDOWN
# ═══════════════════════════════════════════════

def markdown_to_html(md_text: str) -> str:
    """Convert Markdown to HTML."""
    import markdown2
    return markdown2.markdown(
        md_text,
        extras=["fenced-code-blocks", "tables", "header-ids", "strike", "task_list"],
    )


def html_to_markdown(html: str) -> str:
    """Convert HTML to Markdown."""
    from markdownify import markdownify
    return markdownify(html, heading_style="ATX", strip=["img", "script", "style"])


def sanitize_html(html: str, allowed_tags: list[str] | None = None) -> str:
    """Sanitize HTML to prevent XSS. Only allows safe tags."""
    import bleach
    tags = allowed_tags or [
        "p", "br", "strong", "em", "a", "ul", "ol", "li",
        "h1", "h2", "h3", "h4", "blockquote", "code", "pre",
        "table", "thead", "tbody", "tr", "th", "td",
    ]
    return bleach.clean(html, tags=tags, attributes={"a": ["href"]}, strip=True)


def slugify(text: str) -> str:
    """Generate a clean URL slug from text."""
    from slugify import slugify as _slugify
    return _slugify(text, max_length=80)


# ═══════════════════════════════════════════════
# VALIDATION
# ═══════════════════════════════════════════════

def validate_email(email: str) -> dict[str, Any]:
    """Validate an email address. Returns normalized email or error."""
    try:
        from email_validator import validate_email as _validate, EmailNotValidError
        result = _validate(email, check_deliverability=False)
        return {
            "valid": True,
            "normalized": result.normalized,
            "domain": result.domain,
        }
    except Exception as e:
        return {"valid": False, "error": str(e), "normalized": email}


def validate_url(url: str) -> bool:
    """Check if a URL is valid."""
    import validators
    return bool(validators.url(url))


def validate_domain(domain: str) -> bool:
    """Check if a domain name is valid."""
    import validators
    return bool(validators.domain(domain))


# ═══════════════════════════════════════════════
# IMAGE PROCESSING
# ═══════════════════════════════════════════════

def resize_image(
    input_path: str | Path, output_path: str | Path,
    width: int, height: int, quality: int = 85,
) -> bool:
    """Resize an image to specified dimensions."""
    try:
        from PIL import Image
        img = Image.open(str(input_path))
        img = img.resize((width, height), Image.Resampling.LANCZOS)
        img.save(str(output_path), quality=quality, optimize=True)
        return True
    except Exception:
        return False


def create_thumbnail(
    input_path: str | Path, output_path: str | Path,
    size: tuple[int, int] = (300, 300),
) -> bool:
    """Create a thumbnail from an image."""
    try:
        from PIL import Image
        img = Image.open(str(input_path))
        img.thumbnail(size, Image.Resampling.LANCZOS)
        img.save(str(output_path), quality=80, optimize=True)
        return True
    except Exception:
        return False


def add_watermark(
    input_path: str | Path, output_path: str | Path,
    text: str = "ARCANA AI", opacity: int = 128,
) -> bool:
    """Add a text watermark to an image."""
    try:
        from PIL import Image, ImageDraw, ImageFont
        img = Image.open(str(input_path)).convert("RGBA")
        overlay = Image.new("RGBA", img.size, (0, 0, 0, 0))
        draw = ImageDraw.Draw(overlay)

        try:
            font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 24)
        except OSError:
            font = ImageFont.load_default()

        bbox = draw.textbbox((0, 0), text, font=font)
        text_w, text_h = bbox[2] - bbox[0], bbox[3] - bbox[1]
        x = img.width - text_w - 10
        y = img.height - text_h - 10
        draw.text((x, y), text, font=font, fill=(255, 255, 255, opacity))

        result = Image.alpha_composite(img, overlay).convert("RGB")
        result.save(str(output_path), quality=85)
        return True
    except Exception:
        return False


# Platform-specific image sizes
IMAGE_SIZES = {
    "instagram_square": (1080, 1080),
    "instagram_portrait": (1080, 1350),
    "instagram_story": (1080, 1920),
    "twitter_post": (1200, 675),
    "linkedin_post": (1200, 627),
    "facebook_post": (1200, 630),
    "youtube_thumbnail": (1280, 720),
    "tiktok_cover": (1080, 1920),
    "og_image": (1200, 630),
}


# ═══════════════════════════════════════════════
# CACHING
# ═══════════════════════════════════════════════

_memory_caches: dict[str, Any] = {}


def get_cache(name: str = "default", maxsize: int = 256, ttl: int = 3600) -> Any:
    """Get or create a TTL cache. Items expire after ttl seconds."""
    from cachetools import TTLCache
    if name not in _memory_caches:
        _memory_caches[name] = TTLCache(maxsize=maxsize, ttl=ttl)
    return _memory_caches[name]


def cached_get(cache_name: str, key: str) -> Any | None:
    """Get a value from a named cache."""
    cache = _memory_caches.get(cache_name)
    if cache is None:
        return None
    return cache.get(key)


def cached_set(cache_name: str, key: str, value: Any, maxsize: int = 256, ttl: int = 3600) -> None:
    """Set a value in a named cache (auto-creates if needed)."""
    cache = get_cache(cache_name, maxsize, ttl)
    cache[key] = value


def get_disk_cache(name: str = "arcana") -> Any:
    """Get a persistent disk-based cache."""
    from diskcache import Cache
    cache_dir = Path("data") / "cache" / name
    cache_dir.mkdir(parents=True, exist_ok=True)
    return Cache(str(cache_dir))


# ═══════════════════════════════════════════════
# FORMATTING
# ═══════════════════════════════════════════════

def human_number(n: float) -> str:
    """Format number for humans: 1234567 → '1.2M'."""
    import humanize
    return humanize.intword(int(n))


def human_money(amount: float) -> str:
    """Format money: 1234.5 → '$1,234.50'."""
    return f"${amount:,.2f}"


def human_time_ago(dt: datetime) -> str:
    """Format datetime as 'X ago': '3 hours ago'."""
    import humanize
    return humanize.naturaltime(dt)


def human_filesize(size_bytes: int) -> str:
    """Format file size: 1048576 → '1.0 MB'."""
    import humanize
    return humanize.naturalsize(size_bytes)


def format_table(data: list[dict[str, Any]], headers: str = "keys") -> str:
    """Format data as a pretty table."""
    from tabulate import tabulate as _tabulate
    return _tabulate(data, headers=headers, tablefmt="pipe")


# ═══════════════════════════════════════════════
# TEMPLATES
# ═══════════════════════════════════════════════

_jinja_env = None


def render_template(template_str: str, **kwargs: Any) -> str:
    """Render a Jinja2 template string with variables."""
    global _jinja_env
    if _jinja_env is None:
        from jinja2 import Environment
        _jinja_env = Environment(autoescape=True)
    template = _jinja_env.from_string(template_str)
    return template.render(**kwargs)


# ═══════════════════════════════════════════════
# JSON (FAST)
# ═══════════════════════════════════════════════

def fast_json_dumps(obj: Any) -> str:
    """Fast JSON serialization using orjson."""
    import orjson
    return orjson.dumps(obj, default=str).decode()


def fast_json_loads(data: str | bytes) -> Any:
    """Fast JSON deserialization using orjson."""
    import orjson
    return orjson.loads(data)


# ═══════════════════════════════════════════════
# DATETIME
# ═══════════════════════════════════════════════

def now_utc() -> datetime:
    """Get current UTC datetime."""
    return datetime.now(timezone.utc)


def parse_date(text: str) -> datetime | None:
    """Parse a date from any format."""
    import arrow
    try:
        return arrow.get(text).datetime
    except Exception:
        return None


def time_until(dt: datetime) -> str:
    """Human-readable time until a datetime."""
    import arrow
    return arrow.get(dt).humanize()


# ═══════════════════════════════════════════════
# METRICS (Prometheus)
# ═══════════════════════════════════════════════

_metrics: dict[str, Any] = {}


def counter(name: str, description: str = "") -> Any:
    """Get or create a Prometheus counter."""
    from prometheus_client import Counter
    if name not in _metrics:
        _metrics[name] = Counter(name, description or name)
    return _metrics[name]


def histogram(name: str, description: str = "", buckets: tuple = ()) -> Any:
    """Get or create a Prometheus histogram."""
    from prometheus_client import Histogram
    if name not in _metrics:
        _metrics[name] = Histogram(name, description or name, buckets=buckets or Histogram.DEFAULT_BUCKETS)
    return _metrics[name]


def gauge(name: str, description: str = "") -> Any:
    """Get or create a Prometheus gauge."""
    from prometheus_client import Gauge
    if name not in _metrics:
        _metrics[name] = Gauge(name, description or name)
    return _metrics[name]


# ═══════════════════════════════════════════════
# TOKEN COUNTING
# ═══════════════════════════════════════════════

def count_tokens(text: str, model: str = "cl100k_base") -> int:
    """Count tokens for cost estimation."""
    try:
        import tiktoken
        enc = tiktoken.get_encoding(model)
        return len(enc.encode(text))
    except Exception:
        return len(text) // 4  # Fallback approximation


# ═══════════════════════════════════════════════
# URL PROCESSING
# ═══════════════════════════════════════════════

def build_url(base: str, **params: Any) -> str:
    """Build a URL with query parameters."""
    from furl import furl
    f = furl(base)
    f.args.update(params)
    return str(f)


def extract_domain(url: str) -> str:
    """Extract domain from a URL."""
    from furl import furl
    try:
        return furl(url).host or ""
    except Exception:
        return ""


# ═══════════════════════════════════════════════
# PDF GENERATION
# ═══════════════════════════════════════════════

def generate_pdf(
    title: str,
    content_blocks: list[dict[str, str]],
    output_path: str | Path,
    author: str = "ARCANA AI",
) -> bool:
    """Generate a professional PDF document.

    Args:
        title: Document title
        content_blocks: List of {"type": "heading"|"paragraph"|"bullet", "text": str}
        output_path: Where to save the PDF
        author: Document author metadata
    """
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, ListFlowable, ListItem

        doc = SimpleDocTemplate(
            str(output_path),
            pagesize=letter,
            topMargin=1 * inch,
            bottomMargin=1 * inch,
            leftMargin=1 * inch,
            rightMargin=1 * inch,
        )
        doc.author = author
        doc.title = title

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle", parent=styles["Title"], fontSize=24, spaceAfter=20,
        )
        heading_style = ParagraphStyle(
            "CustomHeading", parent=styles["Heading2"], fontSize=16, spaceAfter=10, spaceBefore=15,
        )
        body_style = styles["BodyText"]

        story = [Paragraph(title, title_style), Spacer(1, 0.3 * inch)]

        bullet_buffer: list[str] = []

        def flush_bullets() -> None:
            nonlocal bullet_buffer
            if bullet_buffer:
                items = [ListItem(Paragraph(b, body_style)) for b in bullet_buffer]
                story.append(ListFlowable(items, bulletType="bullet"))
                story.append(Spacer(1, 0.1 * inch))
                bullet_buffer = []

        for block in content_blocks:
            btype = block.get("type", "paragraph")
            text = block.get("text", "")

            if btype != "bullet":
                flush_bullets()

            if btype == "heading":
                story.append(Paragraph(text, heading_style))
            elif btype == "paragraph":
                story.append(Paragraph(text, body_style))
                story.append(Spacer(1, 0.1 * inch))
            elif btype == "bullet":
                bullet_buffer.append(text)

        flush_bullets()
        doc.build(story)
        return True
    except Exception:
        return False


def generate_invoice_pdf(
    output_path: str | Path,
    invoice_number: str,
    client_name: str,
    items: list[dict[str, Any]],
    total: float,
    date: str = "",
) -> bool:
    """Generate a simple invoice PDF.

    Args:
        items: List of {"description": str, "quantity": int, "price": float}
    """
    if not date:
        date = now_utc().strftime("%Y-%m-%d")

    blocks = [
        {"type": "heading", "text": f"Invoice #{invoice_number}"},
        {"type": "paragraph", "text": f"Date: {date}"},
        {"type": "paragraph", "text": f"Bill To: {client_name}"},
        {"type": "heading", "text": "Items"},
    ]
    for item in items:
        desc = item.get("description", "")
        qty = item.get("quantity", 1)
        price = item.get("price", 0)
        blocks.append({"type": "bullet", "text": f"{desc} — {qty}x @ ${price:.2f} = ${qty * price:.2f}"})

    blocks.append({"type": "heading", "text": f"Total: ${total:.2f}"})
    blocks.append({"type": "paragraph", "text": "Thank you for your business! — Arcana Operations LLC"})

    return generate_pdf(f"Invoice {invoice_number}", blocks, output_path)


# ═══════════════════════════════════════════════
# QR CODE GENERATION
# ═══════════════════════════════════════════════

def generate_qr_code(
    data: str, output_path: str | Path,
    size: int = 10, border: int = 2,
) -> bool:
    """Generate a QR code image."""
    try:
        import qrcode
        qr = qrcode.QRCode(version=1, box_size=size, border=border)
        qr.add_data(data)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        img.save(str(output_path))
        return True
    except Exception:
        return False


def generate_micro_qr(data: str, output_path: str | Path) -> bool:
    """Generate a compact QR code using segno (supports Micro QR)."""
    try:
        import segno
        qr = segno.make(data)
        qr.save(str(output_path), scale=8, border=2)
        return True
    except Exception:
        return False


# ═══════════════════════════════════════════════
# EXCEL GENERATION
# ═══════════════════════════════════════════════

def generate_excel(
    data: list[dict[str, Any]],
    output_path: str | Path,
    sheet_name: str = "Report",
) -> bool:
    """Generate an Excel file from a list of dicts."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        if not data:
            wb.save(str(output_path))
            return True

        # Header row (bold, blue background)
        headers = list(data[0].keys())
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Data rows
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))

        # Auto-fit column widths
        for col_idx, header in enumerate(headers, 1):
            max_len = max(len(str(header)), *(len(str(row.get(header, ""))) for row in data))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 50)

        wb.save(str(output_path))
        return True
    except Exception:
        return False


# ═══════════════════════════════════════════════
# FAST DATAFRAMES (Polars)
# ═══════════════════════════════════════════════

def dataframe_from_dicts(data: list[dict[str, Any]]) -> Any:
    """Create a Polars DataFrame from a list of dicts."""
    import polars as pl
    return pl.DataFrame(data)


def dataframe_summary(data: list[dict[str, Any]], group_by: str, agg_col: str, agg: str = "sum") -> list[dict]:
    """Aggregate data using Polars. Returns list of dicts.

    Args:
        agg: "sum", "mean", "count", "min", "max"
    """
    import polars as pl
    df = pl.DataFrame(data)
    agg_fn = getattr(pl.col(agg_col), agg)
    result = df.group_by(group_by).agg(agg_fn().alias(f"{agg_col}_{agg}"))
    return result.to_dicts()


# ═══════════════════════════════════════════════
# SIGNED TOKENS (itsdangerous)
# ═══════════════════════════════════════════════

_signer_secret = None


def _get_signer_secret() -> str:
    """Get signing secret from env or generate one."""
    global _signer_secret
    if _signer_secret is None:
        import os
        _signer_secret = os.getenv("SIGNING_SECRET", "arcana-default-change-me")
    return _signer_secret


def sign_token(data: str, salt: str = "arcana") -> str:
    """Create a signed token (e.g., for unsubscribe links)."""
    from itsdangerous import URLSafeTimedSerializer
    s = URLSafeTimedSerializer(_get_signer_secret())
    return s.dumps(data, salt=salt)


def verify_token(token: str, salt: str = "arcana", max_age: int = 86400) -> str | None:
    """Verify a signed token. Returns original data or None if expired/invalid.

    Args:
        max_age: Maximum age in seconds (default 24 hours)
    """
    from itsdangerous import URLSafeTimedSerializer, SignatureExpired, BadSignature
    s = URLSafeTimedSerializer(_get_signer_secret())
    try:
        return s.loads(token, salt=salt, max_age=max_age)
    except (SignatureExpired, BadSignature):
        return None


# ═══════════════════════════════════════════════
# UNIFIED NOTIFICATIONS (Apprise)
# ═══════════════════════════════════════════════

def send_notification(
    title: str, body: str,
    urls: list[str] | None = None,
) -> bool:
    """Send notification to multiple channels via Apprise.

    URLs format examples:
        - "discord://webhook_id/webhook_token"
        - "tgram://bot_token/chat_id"
        - "slack://token_a/token_b/token_c/#channel"
        - "mailto://user:pass@gmail.com"
    """
    try:
        import apprise
        ap = apprise.Apprise()
        if urls:
            for url in urls:
                ap.add(url)
        return ap.notify(title=title, body=body)
    except Exception:
        return False


# ═══════════════════════════════════════════════
# GEOCODING (Lead Scoring)
# ═══════════════════════════════════════════════

def geocode_location(location: str) -> dict[str, Any] | None:
    """Geocode a location string to lat/lng + details."""
    try:
        from geopy.geocoders import Nominatim
        geolocator = Nominatim(user_agent="arcana-ai")
        result = geolocator.geocode(location, timeout=5)
        if result:
            return {
                "address": result.address,
                "latitude": result.latitude,
                "longitude": result.longitude,
            }
    except Exception:
        pass
    return None


def distance_miles(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Calculate distance between two points in miles."""
    from geopy.distance import geodesic
    return geodesic((lat1, lon1), (lat2, lon2)).miles


# Portland, OR coordinates for lead scoring
ARCANA_HQ = (45.5152, -122.6784)


def lead_distance_score(location: str) -> float:
    """Score a lead 0-1 based on proximity to Portland (closer = higher)."""
    geo = geocode_location(location)
    if not geo:
        return 0.5  # Unknown location gets neutral score
    dist = distance_miles(
        ARCANA_HQ[0], ARCANA_HQ[1],
        geo["latitude"], geo["longitude"],
    )
    # Within 50 miles = 1.0, within 500 miles = 0.7, beyond = 0.3
    if dist < 50:
        return 1.0
    elif dist < 500:
        return 0.7
    elif dist < 2000:
        return 0.5
    return 0.3


# ═══════════════════════════════════════════════
# DNS / DOMAIN INTELLIGENCE
# ═══════════════════════════════════════════════

def check_mx_records(domain: str) -> list[str]:
    """Check if a domain has MX records (can receive email)."""
    try:
        import dns.resolver
        answers = dns.resolver.resolve(domain, "MX")
        return [str(r.exchange).rstrip(".") for r in answers]
    except Exception:
        return []


def domain_has_email(domain: str) -> bool:
    """Check if a domain can receive email (has MX records)."""
    return len(check_mx_records(domain)) > 0


# ═══════════════════════════════════════════════
# CONTENT VARIANTS (Markov Chains)
# ═══════════════════════════════════════════════

def train_markov_model(corpus: list[str]) -> Any:
    """Train a Markov chain model on a corpus of texts."""
    import markovify
    combined = "\n".join(corpus)
    return markovify.Text(combined, state_size=2)


def generate_variants(model: Any, count: int = 5, max_chars: int = 280) -> list[str]:
    """Generate text variants from a trained Markov model."""
    results = []
    for _ in range(count * 3):  # Over-generate to filter
        sentence = model.make_short_sentence(max_chars)
        if sentence and sentence not in results:
            results.append(sentence)
        if len(results) >= count:
            break
    return results


# ═══════════════════════════════════════════════
# FAST HASHING (xxhash)
# ═══════════════════════════════════════════════

def fast_hash(data: str | bytes) -> str:
    """Ultra-fast content hash using xxhash (10x faster than MD5)."""
    import xxhash
    if isinstance(data, str):
        data = data.encode()
    return xxhash.xxh64(data).hexdigest()


def content_fingerprint(text: str) -> str:
    """Generate a fingerprint for deduplication (normalized + hashed)."""
    import re
    normalized = re.sub(r"\s+", " ", text.lower().strip())
    return fast_hash(normalized)


# ═══════════════════════════════════════════════
# SEMANTIC DIFFING
# ═══════════════════════════════════════════════

def deep_diff(old: dict | list, new: dict | list) -> dict[str, Any]:
    """Get semantic diff between two data structures."""
    from deepdiff import DeepDiff
    diff = DeepDiff(old, new, ignore_order=True)
    return dict(diff) if diff else {}


def has_changed(old: dict | list, new: dict | list) -> bool:
    """Check if two data structures differ."""
    return bool(deep_diff(old, new))


# ═══════════════════════════════════════════════
# SITEMAP PARSING (Intel)
# ═══════════════════════════════════════════════

def parse_sitemap(sitemap_url: str) -> list[str]:
    """Parse a sitemap and return all page URLs."""
    try:
        from usp.tree import sitemap_tree_for_homepage
        tree = sitemap_tree_for_homepage(sitemap_url)
        return [page.url for page in tree.all_pages() if page.url]
    except Exception:
        return []


# ═══════════════════════════════════════════════
# GRAPH ANALYSIS (Networks)
# ═══════════════════════════════════════════════

def build_referral_graph(
    edges: list[tuple[str, str, float]],
) -> Any:
    """Build a weighted graph from (source, target, weight) edges."""
    import networkx as nx
    G = nx.DiGraph()
    for source, target, weight in edges:
        G.add_edge(source, target, weight=weight)
    return G


def top_influencers(graph: Any, top_n: int = 10) -> list[tuple[str, float]]:
    """Find top influencers in a network by PageRank."""
    import networkx as nx
    pr = nx.pagerank(graph)
    sorted_nodes = sorted(pr.items(), key=lambda x: -x[1])
    return sorted_nodes[:top_n]


# ═══════════════════════════════════════════════
# COLOR EXTRACTION
# ═══════════════════════════════════════════════

def extract_colors(image_path: str | Path, count: int = 5) -> list[tuple[int, int, int]]:
    """Extract dominant colors from an image as RGB tuples."""
    try:
        from colorthief import ColorThief
        ct = ColorThief(str(image_path))
        return ct.get_palette(color_count=count, quality=10)
    except Exception:
        return []


def dominant_color(image_path: str | Path) -> tuple[int, int, int] | None:
    """Get the single most dominant color from an image."""
    try:
        from colorthief import ColorThief
        ct = ColorThief(str(image_path))
        return ct.get_color(quality=10)
    except Exception:
        return None


def rgb_to_hex(rgb: tuple[int, int, int]) -> str:
    """Convert RGB tuple to hex color string."""
    return "#{:02x}{:02x}{:02x}".format(*rgb)
